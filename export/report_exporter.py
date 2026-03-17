"""Report export to CSV and Excel."""

import csv
import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ReportExporter:
    """Exports report data to CSV or Excel."""

    COLUMNS = [
        "test_name", "module", "run_id", "run_by_who", "status",
        "duration", "start_time", "end_time", "error_message",
    ]

    def to_csv(self, tests: List[Dict[str, Any]]) -> bytes:
        """Export tests to CSV."""
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(self.COLUMNS)
        for t in tests:
            w.writerow([t.get(c) if t.get(c) is not None else "" for c in self.COLUMNS])
        return out.getvalue().encode("utf-8-sig")

    def to_excel(
        self,
        tests: List[Dict[str, Any]],
        run_id: str = "",
        summary: Optional[Dict[str, Any]] = None,
        modules: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        """Export tests to .xlsx."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Tests"

        for i, col in enumerate(self.COLUMNS, 1):
            ws.cell(row=1, column=i, value=col.replace("_", " ").title())
        for ri, t in enumerate(tests, 2):
            for ci, col in enumerate(self.COLUMNS, 1):
                v = t.get(col, "")
                if v is not None:
                    ws.cell(row=ri, column=ci, value=v)
        for i in range(1, len(self.COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        if summary and run_id:
            s = wb.create_sheet("Summary", 0)
            s["A1"], s["A2"] = "Run Report Summary", f"Run ID: {run_id}"
            s["A4"], s["B4"] = "Total Tests", summary.get("total_tests", 0)
            s["A5"], s["B5"] = "Passed", summary.get("passed", 0)
            s["A6"], s["B6"] = "Failed", summary.get("failed", 0)
            s["A7"], s["B7"] = "Duration (s)", summary.get("duration", 0)

        if modules:
            m = wb.create_sheet("Modules", 1 if summary else 0)
            cols = ["module", "passed", "failed", "total", "duration"]
            for i, c in enumerate(cols, 1):
                m.cell(row=1, column=i, value=c.replace("_", " ").title())
            for ri, row in enumerate(modules, 2):
                for i, c in enumerate(cols, 1):
                    v = row.get(c, "")
                    if v is not None:
                        m.cell(row=ri, column=i, value=v)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def run_report_to_csv(self, report: Dict[str, Any]) -> bytes:
        """Export run report to CSV."""
        return self.to_csv(report.get("tests", []))

    def run_report_to_excel(self, report: Dict[str, Any]) -> bytes:
        """Export run report to Excel."""
        tests = report.get("tests", [])
        return self.to_excel(tests, report.get("run_id", ""), report.get("summary"), report.get("modules"))
